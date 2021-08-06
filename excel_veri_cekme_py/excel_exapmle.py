from openpyxl import Workbook,load_workbook
def yeni():
    wb = Workbook()

    ws = wb.active # Aktif olan çalışma alanı

    ws.title = "İlk Çalışma Alanı" # Aktif olan ilk çalışma alanı ismini değiştiriyoruz.
    ws = wb.create_sheet("Posta Kodları")
    ws = wb.create_sheet("Ülkeler")

    ws = wb["Posta Kodları"] # Çalışma alanı değiştiriyoruz.

    ws['A15'] = 42
    ws['B25'] = "Merhaba"
    ws.append([1, 2, 3, "Merhaba", "Dünya"])  # Sıradaki satıra sırasıyla dizi elemanlarını ekler

    print(wb.sheetnames) 
    print(ws)

    wb.save("sözlük.xlsx")

def olan():
    wb =load_workbook("sözlük.xlsx")
    ws = wb.active
    print(ws)
    
    ws['A10'] = 42
    ws['B20'] = "Merhaba"
    ws.append([1, 2, 3, "Merhaba", "Dünya"])  # Sıradaki satıra sırasıyla dizi elemanlarını ekler
    ws['A1'] = 42
    ws.append([1, 2])
    wb.save("sözlük.xlsx") # En sonunda bunu kapatıyoruz.

olan()

# Önceki yazılanları siliyor !!!